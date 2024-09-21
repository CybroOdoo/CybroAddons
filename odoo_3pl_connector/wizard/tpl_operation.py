# -*- coding: utf-8 -*-
################################################################################
#
#    Cybrosys Technologies Pvt. Ltd.
#
#    Copyright (C) 2024-TODAY Cybrosys Technologies(<https://www.cybrosys.com>).
#    Author: Unnimaya C O (odoo@cybrosys.com)
#
#    You can modify it under the terms of the GNU AFFERO
#    GENERAL PUBLIC LICENSE (AGPL v3), Version 3.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU AFFERO GENERAL PUBLIC LICENSE (AGPL v3) for more details.
#
#    You should have received a copy of the GNU AFFERO GENERAL PUBLIC LICENSE
#    (AGPL v3) along with this program.
#    If not, see <http://www.gnu.org/licenses/>.
#
################################################################################
import io
import json
import os
import openpyxl
import xlsxwriter
from ftplib import FTP
from odoo import api, fields, models, _
from odoo.exceptions import ValidationError
from odoo.tools import date_utils


class TplOperation(models.TransientModel):
    """Model for performing the tpl operations"""
    _name = 'tpl.operation'
    _description = 'Tpl Operation'
    _rec_name = 'warehouse_id'

    export = fields.Selection([
        ('sales', 'Sales'),
        ('return', 'Sales Return'),
        ('purchase', 'Purchase')],
        string='Items to Export/ Import',
        default='sales', required=True,
        help='Select the items to import/export')
    warehouse_id = fields.Many2one('stock.warehouse',
                                   string="Warehouse",
                                   help="Select the warehouse for the "
                                        "operations", required=1)
    operation = fields.Selection([('export', 'Export'),
                                  ('import', 'Import')],
                                 default='import',
                                 string='Operation', help='The operation which '
                                                          'you want to perform')
    picking_ids = fields.Many2many('stock.picking',
                                   string='Sales Picking',
                                   help='All the sales pickings which are not '
                                        'exported yet.')

    @api.onchange('export')
    def _onchange_export(self):
        """Method for calling bind_data() while changing the value of field
         export"""
        self.bind_data()

    def bind_data(self):
        """Method for binding data to the one2many field"""
        if self.operation == 'export':
            if self.export == 'sales':
                domain = 'WH-DELIVERY'
            elif self.export == 'return':
                domain = 'WH-RETURNS'
            else:
                domain = 'WH-RECEIPTS'
            self.picking_ids = (
                self.env['stock.picking'].search([('state', '=', 'assigned'),
                                                  ('is_exported', '=', False)]).
                filtered(
                    lambda x: x.picking_type_id.barcode == domain))

    def action_execute(self):
        """Execute button action"""
        if self.operation == 'export' and self.picking_ids:
            if not self.warehouse_id.server_id:
                raise ValidationError(_('3PL not configured for selected '
                                        'warehouse.'))
            picking_list = self.env['stock.move.line'].search_read(
                domain=[('picking_id', 'in', self.picking_ids.ids)],
                fields=[
                    'picking_id',
                    'product_id',
                    'reserved_qty',
                    'picking_partner_id',
                ])
            for item in picking_list:
                if item['picking_partner_id']:
                    partner = self.env['res.partner'].search_read(
                        domain=[('id', '=', item['picking_partner_id'][0])],
                        fields=[
                            'street',
                            'street2',
                            'zip',
                            'city',
                            'state_id',
                            'country_id',
                            'email',
                            'phone',
                            'mobile'
                        ])
                    for rec in partner[0].keys():
                        if rec != 'id':
                            item[rec] = partner[0][rec]
            transfer_dict = {'sales': 'Export_Sales',
                             'return': 'Export_Sales_Return',
                             'purchase': 'Export_Purchase'}
            return {
                'type': 'ir.actions.report',
                'data': {
                    'model': 'tpl.operation',
                    'output_format': 'xlsx',
                    'report_name': 'Excel Report',
                    'options': json.dumps({
                        'warehouse_id': self.warehouse_id.id,
                        'picking_list': picking_list,
                        'transfer': transfer_dict[self.export]
                    }, default=date_utils.json_default)
                },
                'report_type': 'xlsx',
            }
        if self.operation == 'import':
            ftp = FTP(self.warehouse_id.server_id.host,
                      self.warehouse_id.server_id.username,
                      self.warehouse_id.server_id.password)
            try:
                ftp.encoding = "utf-8"
                transfer_dict = {'sales': 'Import_Sales',
                                 'return': 'Import_Sales_Return',
                                 'purchase': 'Import_Purchase'}
                local_filepath = os.path.normpath(
                    os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                 "..", "demo",
                                 "imported_report.xlsx")
                )
                ftp.cwd('/' + transfer_dict[self.export])
                for remote_file in ftp.nlst():
                    with open(local_filepath, 'wb') as local_file:
                        ftp.retrbinary(f"RETR {remote_file}",
                                       local_file.write)
                    workbook = openpyxl.load_workbook(local_filepath)
                    sheet = workbook['Sheet1']
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        picking = self.env['stock.picking'].search(
                            [('name', '=', row[0])])
                        if picking:
                            picking.action_confirm()
                            picking.action_assign()
                            picking.move_ids._set_quantities_to_reservation()
                            picking.button_validate()
                            picking.sudo().write({
                                'state':'done',
                                'is_delivered': True
                            })
                    ftp.delete(remote_file)
                self.bind_data()
                return {'type': 'ir.actions.act_window_close'}
            except Exception as e:
                raise ValidationError(f"An Error occurred: {str(e)}")
            finally:
                ftp.quit()

    def get_xlsx_report(self, data, response):
        """Organizing xlsx report"""
        data = json.loads(data)
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        sheet = workbook.add_worksheet()
        cell_format = workbook.add_format(
            {'font_size': '14', 'bold': True, 'align': 'center',
             'valign': 'center', 'border': True})
        body = workbook.add_format(
            {'align': 'left', 'text_wrap': True, 'border': True})
        body_right = workbook.add_format(
            {'align': 'right', 'text_wrap': True, 'border': True})
        sheet.set_column('A1:A2', 18)
        sheet.set_column('B1:B2', 50)
        sheet.set_column('C1:C2', 12)
        sheet.set_column('D1:M1', 18)
        sheet.set_row(0, 20)
        sheet.set_row(1, 20)
        sheet.merge_range('A1:A2', 'Reference', cell_format)
        sheet.merge_range('B1:B2', 'Product', cell_format)
        sheet.merge_range('C1:C2', 'Quantity', cell_format)
        sheet.write('D2', 'Delivery Name', cell_format)
        sheet.merge_range('D1:M1', 'Delivery', cell_format)
        sheet.write('E2', 'Street', cell_format)
        sheet.write('F2', 'Street2', cell_format)
        sheet.write('G2', 'City', cell_format)
        sheet.write('H2', 'State', cell_format)
        sheet.write('I2', 'Country', cell_format)
        sheet.write('J2', 'Zip', cell_format)
        sheet.write('K2', 'Email', cell_format)
        sheet.write('L2', 'Phone', cell_format)
        sheet.write('M2', 'Mobile', cell_format)
        row = 2
        column = 0
        for item in data['picking_list']:
            sheet.write(row, column, item['picking_id'][1], body)
            sheet.write(row, column + 1, item['product_id'][1], body)
            sheet.write(row, column + 2,
                        "{:.2f}".format(int(item['reserved_qty'])), body_right)
            sheet.write(row, column + 3, item['picking_partner_id'][1],
                        body) if item['picking_partner_id'] else (
                sheet.write(row, column + 3, " ", body))
            sheet.write(row, column + 4, item['street'],
                        body) if 'street' in item.keys() and item['street'] \
                else sheet.write(row, column + 4, " ", body)
            sheet.write(row, column + 5, item['street2'],
                        body) if 'street2' in item.keys() and item['street2'] \
                else sheet.write(row, column + 5, " ", body)
            sheet.write(row, column + 6, item['city'],
                        body) if 'city' in item.keys() and item['city'] else (
                sheet.write(row, column + 6, " ", body))
            sheet.write(row, column + 7, item['state_id'][1],
                        body) if ('state_id' in item.keys() and
                                  item['state_id']) else (
                sheet.write(row, column + 7, " ", body))
            sheet.write(row, column + 8, item['country_id'][1],
                        body) if ('country_id' in item.keys() and
                                  item['country_id']) else (
                sheet.write(row, column + 8, " ", body))
            sheet.write(row, column + 9, item['zip'],
                        body) if ('zip' in item.keys() and item['zip']) else (
                sheet.write(row, column + 9, " ", body))
            sheet.write(row, column + 10, item['email'],
                        body) if ('email' in item.keys() and item['email']) \
                else sheet.write(row, column + 10, " ", body)
            sheet.write(row, column + 11, item['phone'],
                        body) if ('phone' in item.keys() and item['phone']) \
                else sheet.write(row, column + 11, " ", body)
            sheet.write(row, column + 12, item['mobile'],
                        body) if ('mobile' in item.keys() and item['mobile']) \
                else sheet.write(row, column + 12, " ", body)
            row = row + 1
        workbook.close()
        remote_file = (data['transfer'] + '/exported_report_' +
                       fields.Datetime.now(
                       ).strftime("%Y_%m_%d_%H_%M_%S_%f")[:-3] + '.xlsx')
        try:
            server = self.env['stock.warehouse'].browse(
                int(data['warehouse_id'])).server_id
            if not server:
                ValidationError(_('3PL not configured for the Warehouse'))
            ftp = FTP(server.host, server.username,
                      server.password)
            ftp.encoding = "utf-8"
            local_filepath = os.path.normpath(os.path.join(
                os.path.dirname(os.path.abspath(__file__)), "..", "demo",
                "imported_report.xlsx"))
            with open(local_filepath, 'wb') as temp_file:
                temp_file.write(output.getvalue())
            with open(local_filepath, 'rb') as file:
                ftp.storbinary(f'STOR {remote_file}', file)
            output.seek(0)
            response.stream.write(output.read())
            output.close()
            for item in data['picking_list']:
                self.env['stock.picking'].browse(
                    item['picking_id'][0]).write({'is_exported': True})
        except Exception as e:
            raise ValidationError(f"An error occurred: {e}")
