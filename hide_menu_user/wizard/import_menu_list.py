# -*- coding: utf-8 -*-
#############################################################################
#
#    Cybrosys Technologies Pvt. Ltd.
#
#    Copyright (C) 2022-TODAY Cybrosys Technologies(<https://www.cybrosys.com>)
#    Author: Cybrosys Techno Solutions(<https://www.cybrosys.com>)
#
#    You can modify it under the terms of the GNU LESSER
#    GENERAL PUBLIC LICENSE (LGPL v3), Version 3.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU LESSER GENERAL PUBLIC LICENSE (LGPL v3) for more details.
#
#    You should have received a copy of the GNU LESSER GENERAL PUBLIC LICENSE
#    (LGPL v3) along with this program.
#    If not, see <http://www.gnu.org/licenses/>.
#
#############################################################################

import base64
import openpyxl
from io import BytesIO
from odoo import models, fields, _
from odoo.exceptions import UserError


class OrderLinesWizard(models.TransientModel):
    """ This class contain the fields of wizard and it will import the xslx file"""
    _name = 'import.menu.list.wizard'
    _description = "Import the menu list"

    name = fields.Char(string="Name", Readonly=True)
    import_file = fields.Binary(string="Upload Files", help="Select the xlsx file to import ")
    user_id = fields.Integer(string="User Id")

    def action_import_xls(self):
        """This function will help to import the Excel sheet"""
        try:
            wb = openpyxl.load_workbook(
                filename=BytesIO(base64.b64decode(self.import_file)), read_only=True
            )
            ws = wb.active
        except:
            raise UserError(
                _('Please insert a valid file'))

        col = -1
        flag = 0
        for rec in ws.iter_rows(min_row=1, max_row=1, min_col=None, max_col=None, values_only=True):
            for head in rec:
                col = col + 1
                if head == 'Parent Path':
                    flag = 1
                    break
            if flag == 0:
                raise UserError(
                    _('Please insert a parent path in the file'))
        for record in ws.iter_rows(min_row=2, max_row=None, min_col=None, max_col=None, values_only=True):
            menu_exist = self.env['ir.ui.menu'].search([('parent_path', '=', record[col])]).id
            if not menu_exist:
                raise UserError(
                    _('Please have valid parent path'))
            res = {
                'res_users_id': self.user_id,
                'ir_ui_menu_id': menu_exist,
            }
            self.env['res.users'].search([('id', '=', res['res_users_id'])]).write(
                {'hide_menu_ids':  [(4, res['ir_ui_menu_id'])]})
