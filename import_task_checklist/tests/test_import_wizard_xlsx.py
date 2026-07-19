# -*- coding: utf-8 -*-
###############################################################################
#
#    Cybrosys Technologies Pvt. Ltd.
#
#    Copyright (C) 2026-TODAY Cybrosys Technologies(<https://www.cybrosys.com>)
#    Author: Cybrosys Techno Solutions(<https://www.cybrosys.com>)
#
#    You can modify it under the terms of the GNU AFFERO
#    GENERAL PUBLIC LICENSE (AGPL v3), Version 3.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU AFFERO GENERAL PUBLIC LICENSE (AGPL v3) for more details.
#
#    You should have received a copy of the GNU AFFERO GENERAL PUBLIC LICENSE
#    (AGPL v3) along with this program.
#    If not, see <http://www.gnu.org/licenses/>.
#
###############################################################################

import base64
from io import BytesIO

from odoo.exceptions import UserError
from odoo.tests.common import TransactionCase
from openpyxl import Workbook


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

HEADER = ['Sequence', 'Description', 'Name']


def _xlsx_b64(*rows):
    """Build an openpyxl workbook from row-lists and return base64 bytes."""
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue())


class TestImportWizardXlsx(TransactionCase):
    """Tests for XLSX import — action_import_task_checklist_xlsx()."""

    # ------------------------------------------------------------------
    # Fixtures
    # ------------------------------------------------------------------
    def setUp(self):
        super().setUp()
        self.Wizard    = self.env['import.task.checklist']
        self.Checklist = self.env['task.checklist']
        self.company   = self.env.company

    def _wizard(self, file_content, filename='test.xlsx', file_type='xls'):
        return self.Wizard.create({
            'file_type':    file_type,
            'file_content': file_content,
            'filename':     filename,
        })

    # ------------------------------------------------------------------
    # Happy path — column layout verification
    # ------------------------------------------------------------------
    def test_xlsx_name_comes_from_col2(self):
        """Record 'name' must be taken from column C / index 2 (v17 layout)."""
        content = _xlsx_b64(HEADER, ['1', 'XL Desc', 'XL Task Name'])
        self._wizard(content).action_import_task_checklist_xlsx()
        rec = self.Checklist.search([('name', '=', 'XL Task Name')])
        self.assertTrue(rec, "name should be created from col 2.")

    def test_xlsx_description_comes_from_col1(self):
        """Record 'description' must be taken from column B / index 1 (v17 layout)."""
        content = _xlsx_b64(HEADER, ['1', 'XL Description', 'XL Task'])
        self._wizard(content).action_import_task_checklist_xlsx()
        rec = self.Checklist.search([('name', '=', 'XL Task')])
        self.assertTrue(rec)
        self.assertEqual(
            rec.description, 'XL Description',
            "description should come from col 1."
        )

    def test_xlsx_imports_multiple_rows(self):
        """All data rows (row index >= 1 / min_row=2) should produce separate records."""
        content = _xlsx_b64(
            HEADER,
            ['1', 'D1', 'Task Alpha'],
            ['2', 'D2', 'Task Beta'],
            ['3', 'D3', 'Task Gamma'],
        )
        self._wizard(content).action_import_task_checklist_xlsx()
        for name in ['Task Alpha', 'Task Beta', 'Task Gamma']:
            self.assertTrue(
                self.Checklist.search([('name', '=', name)]),
                f"'{name}' should have been imported."
            )

    # ------------------------------------------------------------------
    # Return value contract
    # ------------------------------------------------------------------
    def test_xlsx_returns_client_action_dict(self):
        """Method must return a dict with type=ir.actions.client."""
        content = _xlsx_b64(HEADER, ['1', 'Ret D', 'Ret Task'])
        result  = self._wizard(content).action_import_task_checklist_xlsx()
        self.assertIsInstance(result, dict)
        self.assertEqual(result.get('type'), 'ir.actions.client')
        self.assertEqual(result.get('tag'),  'display_notification')

    def test_xlsx_notification_type_success(self):
        """params.type must equal 'success'."""
        content = _xlsx_b64(HEADER, ['1', 'S D', 'S Task'])
        result  = self._wizard(content).action_import_task_checklist_xlsx()
        self.assertEqual(result['params']['type'], 'success')

    def test_xlsx_notification_not_sticky(self):
        """params.sticky must be False."""
        content = _xlsx_b64(HEADER, ['1', 'NS D', 'NS Task'])
        result  = self._wizard(content).action_import_task_checklist_xlsx()
        self.assertFalse(result['params'].get('sticky'))

    # ------------------------------------------------------------------
    # Duplicate prevention (keyed on col 1 — same as v17 CSV logic)
    # ------------------------------------------------------------------
    def test_xlsx_no_duplicate_on_reimport(self):
        """Importing the same file twice must not create a second record."""
        content = _xlsx_b64(HEADER, ['1', 'Uniq Desc', 'Uniq Task'])
        self._wizard(content).action_import_task_checklist_xlsx()
        self._wizard(content).action_import_task_checklist_xlsx()
        matches = self.Checklist.search([('name', '=', 'Uniq Task')])
        self.assertEqual(len(matches), 1, "Duplicate should not be created.")

    def test_xlsx_rows_sharing_name_second_skipped(self):
        """Two rows with identical col-2 (name) — second must be skipped."""
        content = _xlsx_b64(
            HEADER,
            ['1', 'Desc One', 'Same Name'],
            ['2', 'Desc Two', 'Same Name'],  # same name → skip
        )
        self._wizard(content).action_import_task_checklist_xlsx()
        self.assertTrue(
            self.Checklist.search([('name', '=', 'Same Name'), ('description', '=', 'Desc One')]),
            "First row should be created."
        )
        self.assertFalse(
            self.Checklist.search([('name', '=', 'Same Name'), ('description', '=', 'Desc Two')]),
            "Row with duplicate col-2 value should be skipped."
        )

    # ------------------------------------------------------------------
    # Invalid / corrupt file
    # ------------------------------------------------------------------
    def test_xlsx_raises_on_corrupt_bytes(self):
        """Corrupt binary that is not a valid xlsx must raise UserError."""
        corrupt = base64.b64encode(b'not a real excel file!!!')
        with self.assertRaises(UserError):
            self._wizard(corrupt).action_import_task_checklist_xlsx()

    def test_xlsx_raises_on_csv_content_passed_as_xlsx(self):
        """A CSV payload submitted as xlsx must raise UserError."""
        csv_bytes = b'Sequence,Description,Name\n1,D,N\n'
        wrong = base64.b64encode(csv_bytes)
        with self.assertRaises(UserError):
            self._wizard(wrong, filename='wrong.csv').action_import_task_checklist_xlsx()

    # ------------------------------------------------------------------
    # Header-only workbook
    # ------------------------------------------------------------------
    def test_xlsx_header_only_creates_no_records(self):
        """Workbook with only a header row must raise UserError."""
        content      = _xlsx_b64(HEADER)
        with self.assertRaises(UserError):
            self._wizard(content).action_import_task_checklist_xlsx()

    # ------------------------------------------------------------------
    # Multi-sheet workbook
    # ------------------------------------------------------------------
    def test_xlsx_all_sheets_are_processed(self):
        """Every worksheet in the workbook should be iterated and imported."""
        wb  = Workbook()
        ws1 = wb.active
        ws1.title = 'Sheet1'
        ws1.append(HEADER)
        ws1.append(['1', 'S1 Desc', 'Sheet1 Task'])

        ws2 = wb.create_sheet('Sheet2')
        ws2.append(HEADER)
        ws2.append(['1', 'S2 Desc', 'Sheet2 Task'])

        buf = BytesIO()
        wb.save(buf)
        content = base64.b64encode(buf.getvalue())

        self._wizard(content).action_import_task_checklist_xlsx()

        self.assertTrue(
            self.Checklist.search([('name', '=', 'Sheet1 Task')]),
            "Sheet1 record should be imported."
        )
        self.assertTrue(
            self.Checklist.search([('name', '=', 'Sheet2 Task')]),
            "Sheet2 record should be imported."
        )

    # ------------------------------------------------------------------
    # Empty cell in name column
    # ------------------------------------------------------------------
    def test_xlsx_empty_name_col_row_skipped(self):
        """Row where col 2 (name) is None/empty should not produce a record."""
        content = _xlsx_b64(
            HEADER,
            ['1', 'Has Desc', None],      # name cell is empty
            ['2', 'Real Desc', 'Real Task'],
        )
        self._wizard(content).action_import_task_checklist_xlsx()
        real = self.Checklist.search([('name', '=', 'Real Task')])
        self.assertTrue(real, "Row with valid name should still be imported.")

    # ------------------------------------------------------------------
    # openpyxl vs xlrd migration guard
    # ------------------------------------------------------------------
    def test_xlsx_openpyxl_workbook_is_readable(self):
        """Workbook saved by openpyxl must be successfully opened by the wizard.
        This test fails if the v19 migration accidentally kept the xlrd import."""
        content = _xlsx_b64(HEADER, ['1', 'Compat Desc', 'Compat Task'])
        try:
            self._wizard(content).action_import_task_checklist_xlsx()
        except UserError as exc:
            self.fail(
                f"openpyxl-generated workbook raised UserError — "
                f"wizard may still be using xlrd: {exc}"
            )
        rec = self.Checklist.search([('name', '=', 'Compat Task')])
        self.assertTrue(rec, "openpyxl workbook should be importable in v19.")
