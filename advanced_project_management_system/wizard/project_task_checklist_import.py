# -*- coding: utf-8 -*-
#############################################################################
#
#    Cybrosys Technologies Pvt. Ltd.
#
#    Copyright (C) 2026-TODAY Cybrosys Technologies(<https://www.cybrosys.com>)
#    Author: Cybrosys Techno Solutions
#
#    You can modify it under the terms of the GNU LESSER
#    GENERAL PUBLIC LICENSE (LGPL v3), Version 3.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU LESSER GENERAL PUBLIC LICENSE (LGPL v3) for more details.
#
#    You should have received a copy of the GNU LESSER GENERAL PUBLIC LICENSE
#    (LGPL v3) along with this program.
#    If not, see <http://www.gnu.org/licenses/>.
#
#############################################################################
from odoo import models, fields, _
from odoo.exceptions import UserError
from openpyxl import load_workbook
import base64
import io
import csv

class ProjectTaskChecklistImport(models.TransientModel):
    """
    Wizard to import project task checklists from external files
    (Excel/XLSX or CSV).
    """
    _name = "project.task.checklist.import"
    _description = "Task custom checklist import"

    file = fields.Binary(
        string='Upload Your File Here',
        required=True,
        help="For updating the file"
    )
    file_name = fields.Char(string="File Name")

    file_type = fields.Selection(
        [('excel', 'Excel'), ('csv', 'CSV')],
        string="File Type",
        required=True,
        help="For determining the file type"
    )

    company_id = fields.Many2one(
        'res.company',
        string="Company",
        default=lambda self: self.env.company
    )

    download_sample_file = fields.Boolean(
        string="Download Sample File"
    )

    def import_custom_checklist(self):
        """
        Parses the uploaded file (Excel or CSV) and creates corresponding
        'project.task.checklist' records.
        """

        # File type validation
        if not self.file_type:
            raise UserError(_("Please select a file type."))

        if not self.file or not self.file_name:
            raise UserError(_("Please upload a file."))

        file_name = self.file_name.lower()

        # ---------- EXCEL (.xlsx ONLY) ----------
        if self.file_type == 'excel':

            if not file_name.endswith('.xlsx'):
                raise UserError(_("Please upload an Excel (.xlsx) file."))

            try:
                file_data = base64.b64decode(self.file)
                wb = load_workbook(io.BytesIO(file_data), data_only=True)
                sheet = wb.active
            except Exception:
                raise UserError(_("Invalid Excel (.xlsx) file."))

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue

                self.env['project.task.checklist'].create({
                    'name': row[0],
                    'description': row[1],
                    'company_id': self.company_id.id
                })

        # ---------- CSV ----------
        elif self.file_type == 'csv':

            if not file_name.endswith('.csv'):
                raise UserError(_("Please upload a CSV file."))

            try:
                csv_data = base64.b64decode(self.file)
                data_file = io.StringIO(csv_data.decode("utf-8"))
                csv_reader = csv.reader(data_file)
                next(csv_reader, None)  # skip header
            except Exception:
                raise UserError(_("Invalid CSV file."))

            for row in csv_reader:
                if not row or not row[0]:
                    continue

                self.env['project.task.checklist'].create({
                    'name': row[0],
                    'description': row[1],
                    'company_id': self.company_id.id
                })
