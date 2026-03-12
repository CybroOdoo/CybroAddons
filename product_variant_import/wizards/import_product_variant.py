# -*- coding: utf-8 -*-
#############################################################################
#
#    Cybrosys Technologies Pvt. Ltd.
#
#    Copyright (C) 2026-TODAY Cybrosys Technologies(<https://www.cybrosys.com>)
#    Author: Cybrosys Techno Solutions(<https://www.cybrosys.com>)
#
#    You can modify it under the terms of the GNU LESSER
#    GENERAL PUBLIC LICENSE (LGPL v3), Version 3.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU LESSER GENERAL PUBLIC LICENSE (LGPL v3) for more details.
#
#    You should have received a copy of the GNU LESSER GENERAL PUBLIC LICENSE
#    (LGPL v3) along with this program.
#    If not, see <http://www.gnu.org/licenses/>.
#
#############################################################################
import os
import io, requests
from openpyxl import load_workbook
import re
import binascii
import tempfile
from odoo import models, fields, _
import csv
import base64
from odoo.exceptions import UserError


class ImportVariant(models.TransientModel):
    """Wizard for selecting the imported Files"""
    _name = 'import.product.variant'
    _description = "Import Product Variants"

    import_file = fields.Selection(
        [('csv', 'CSV File'), ('excel', 'Excel File')], required=True,
        string="Import FIle", help="Import the files")
    method = fields.Selection([('create', 'Create Product'),
                               ('update', 'Update Product'), (
                                   'update_product',
                                   'Update Product Variant'), ],
                              string="Method", required=True,
                              help="Method for importing/Exporting")
    file = fields.Binary(string="File", required=True,
                         help="The file to upload")

    def action_import_product_variant(self):
        """This is used to import/export the product """
        try:
            global list, detailed, invoicing_type
            link = False
            if self.import_file == 'excel':
                try:
                    file_pointer = tempfile.NamedTemporaryFile(delete=False,
                                                               suffix=".xlsx")
                    file_pointer.write(binascii.a2b_base64(self.file))
                    file_pointer.seek(0)
                    workbook = load_workbook(filename=file_pointer.name)
                    sheet = workbook['Sheet1']
                except:
                    raise UserError(_("File not Valid"))
                for index, row in enumerate(sheet.iter_rows(values_only=True)):
                    if index == 0:
                        continue  # Skip the first row
                        # if rec >= 1:
                    if row:
                        row_vals = row
                        if len(row_vals) < int(24):
                            raise UserError(
                                _("Please ensure that you selected "
                                  "the correct file"))
                        product_category = self.env['product.category'].search(
                            [('complete_name', '=', row_vals[6])]).id
                        if product_category:
                            category = product_category
                        else:
                            category = self.env['product.category'].create({
                                'name': row_vals[6].split('/')[0],
                                'complete_name': row_vals[6]
                            })
                            category = category.id
                        product_uom = self.env['uom.uom'].search(
                            [('name', '=', row_vals[7])]).id
                        if product_uom:
                            uom = product_uom
                        else:
                            raise UserError(_("Invalid uom"))

                        description = row_vals[8]

                        def calculate_tax_value(val):
                            if isinstance(val, float) and val < 1:
                                return val * 100, f"{val * 100}%"
                            return val, val

                        def create_tax_record(name, amount, tax_type):
                            match = re.search(r'\d+', name)
                            amount = int(match.group()) if match else amount
                            return self.env['account.tax'].create({
                                'name': name,
                                'amount': amount,
                                'type_tax_use': tax_type,
                            }).id

                        account_val, account_name = calculate_tax_value(
                            row_vals[9])
                        supp_val, supp_name = calculate_tax_value(row_vals[10])

                        account_tax = self.env['account.tax'].search(
                            [('name', '=', account_name),
                             ('type_tax_use', '=', 'sale')]).id
                        supp_tax = self.env['account.tax'].search(
                            [('name', '=', supp_name),
                             ('type_tax_use', '=', 'purchase')]).id

                        tax = account_tax if account_tax else create_tax_record(
                            account_name, account_val, 'sale')
                        supplier_tax = supp_tax if supp_tax else create_tax_record(
                            supp_name, supp_val, 'purchase')

                        kay_val_dict = dict(
                            self.env['product.template']._fields[
                                'type'].selection)
                        # here 'type' is field name
                        for key, val in kay_val_dict.items():
                            if val == row_vals[5]:
                                detailed = key
                        kay_val_dict = dict(
                            self.env['product.template']._fields[
                                'invoice_policy'].selection)
                        # here 'type' is field name
                        for key, val in kay_val_dict.items():
                            if val == row_vals[12]:
                                invoicing_type = key
                        if "http://" in row_vals[23] or "https://" in row_vals[
                            23]:
                            link = base64.b64encode(
                                requests.get(
                                    row_vals[23].strip()).content).replace(
                                b"\n", b"")
                        elif "/home" in row_vals[23]:
                            if os.path.exists(row_vals[23]):
                                with open(row_vals[23], 'rb') as image_file:
                                    link = base64.b64encode(image_file.read())
                        if not row_vals[2] or not row_vals[18]:
                            raise UserError(_("File Must  Contain Internal "
                                              "Reference or Barcode of the "
                                              "Product"))
                        if self.method == 'update':
                            vals = {
                                'default_code': row_vals[2],
                                'name': row_vals[1],
                                'image_1920': link,
                                'sale_ok': row_vals[3],
                                'purchase_ok': row_vals[4],
                                'type': detailed,
                                'categ_id': category,
                                'uom_id': uom,
                                'description': description,
                                'taxes_id': [tax],
                                'supplier_taxes_id': [supplier_tax],
                                'description_sale': row_vals[11],
                                'invoice_policy': invoicing_type,
                                'list_price': row_vals[13],
                                'standard_price': row_vals[14],
                                'weight': row_vals[19],
                                'volume': row_vals[20],
                            }
                            if link:
                                vals.update({'image_1920': link})
                            product = self.env['product.template'].search(
                                [('barcode', '=', row_vals[18])])
                            if product:
                                product.write(vals)
                            else:
                                product = self.env['product.template'].search(
                                    [('default_code', '=', row_vals[2])])
                                if product:
                                    product.write(vals)
                                else:
                                    raise UserError(
                                        _("Please ensure that product "
                                              "having the"
                                              "contains Internal reference or "
                                              "Barcode to with your file"))
                        else:
                            if self.method == 'update_product':
                                vals = {
                                    'default_code': row_vals[2],
                                    'name': row_vals[1],
                                    'image_1920': link,
                                    'sale_ok': row_vals[3],
                                    'purchase_ok': row_vals[4],
                                    'type': detailed,
                                    'categ_id': category,
                                    'uom_id': uom,
                                    'description': description,
                                    'taxes_id': [tax],
                                    'supplier_taxes_id': [supplier_tax],
                                    'supplier_taxes_id': [(6, 0, [supplier_tax])],
                                    'description_sale': row_vals[11],
                                    'invoice_policy': invoicing_type,
                                    'lst_price': row_vals[13],
                                    'standard_price': row_vals[14],
                                    'weight': row_vals[19],
                                    'volume': row_vals[20],
                                }
                                if link:
                                    vals.update({'image_1920': link})
                                product = self.env['product.product'].search(
                                    [('barcode', '=', row_vals[18])])
                                if product:
                                    product.write(vals)
                                else:
                                    product = self.env[
                                        'product.product'].search(
                                        [('default_code', '=', row_vals[2])])
                                    if product:
                                        product.write(vals)
                                    else:
                                        raise UserError(
                                            _("Please ensure that product "
                                              "having the"
                                              "contains Internal reference or "
                                              "Barcode to with your file."))
                            else:
                                vals = {
                                    'default_code': row_vals[2],
                                    'name': row_vals[1],
                                    'image_1920': link,
                                    'sale_ok': row_vals[3],
                                    'purchase_ok': row_vals[4],
                                    'type': detailed,
                                    'categ_id': category,
                                    'uom_id': uom,
                                    'description': description,
                                    'barcode': row_vals[18],
                                    'taxes_id': [tax],
                                    'supplier_taxes_id': [supplier_tax],
                                    'description_sale': row_vals[11],
                                    'invoice_policy': invoicing_type,
                                    'list_price': row_vals[13],
                                    'standard_price': row_vals[14],
                                    'weight': row_vals[19],
                                    'volume': row_vals[20],

                                }
                                product = self.env['product.template'].create(
                                    vals)
                                product.barcode =  row_vals[18]
                                product.write({'barcode': row_vals[18] })
                        values = []
                        for row_val in row_vals[15].split(','):
                            pr_attribute = self.env['product.attribute'].search(
                                [('name', '=', row_val)]).id
                            if pr_attribute:
                                attribute = pr_attribute
                            else:
                                raise UserError(
                                    _("Please update a valid attribute and "
                                      "values"))
                            values.append({'attribute': attribute})
                            for row in row_vals[16].split(','):
                                attri_values = self.env[
                                    'product.attribute.value'].search(
                                    [('attribute_id', '=', attribute),
                                     ('name', '=', row)]).ids
                                if len(attri_values) != 0:
                                    values.extend({attri_values[0]})
                        variant = {}
                        mylist = []
                        for val in values:
                            if isinstance(val, dict):
                                variant = val
                                variant['attribut_value'] = []
                            else:
                                variant['attribut_value'].extend([val])
                            if variant in mylist:
                                pass
                            else:
                                mylist.append(variant)
                        for lst in mylist:
                            val = {
                                'product_tmpl_id': product.id,
                                'attribute_id': lst['attribute'],
                                'value_ids': lst['attribut_value'],
                            }
                            self.env['product.template.attribute.line'].create(
                                val)
            elif self.import_file == 'csv':
                keys = ['Unique Identifier', 'Name', 'Internal Reference',
                        'Can be sold', 'Can be Purchased', 'Product Type',
                        'Category', 'Unit of Measure',
                        'Description',
                        'Customer Taxes', 'Vendor Taxes',
                        'Description for customers', 'Invoicing Policy',
                        'Sales Price', 'Cost', 'Variant Attributes',
                        'Attribute Values', 'Internal Reference', 'Barcode',
                        'Weight', 'Volume', 'Qty On hand',
                        'Responsible', 'image', 'Char', 'Many2many', 'Many2one',
                        'Integer']
                try:
                    files = base64.b64decode(self.file)
                    data = io.StringIO(files.decode("utf-8"))
                    data.seek(0)
                    file_reader = []
                    csv_reader = csv.reader(data, delimiter=',')
                    file_reader.extend(csv_reader)
                except:
                    raise UserError(_("File not Valid"))
                for file in range(len(file_reader)):
                    field = list(map(str, file_reader[file]))
                    values = dict(zip(keys, field))
                    if file >= 1:
                        pro_categ = self.env['product.category'].search(
                            [('complete_name', '=', values['Category'])]).id
                        if pro_categ:
                            pro_category = pro_categ
                        else:
                            category = self.env['product.category'].create({
                                'name': values['Category']
                            })
                            pro_category = category.id
                        unit_uom = self.env['uom.uom'].search(
                            [('name', '=', values['Unit of Measure'])]).id
                        if unit_uom:
                            uom = unit_uom
                        else:
                            raise UserError(_("Invalid uom"))
                        def calculate_tax_value(val):
                            if isinstance(val, float) and val < 1:
                                return val * 100, f"{val * 100}%"
                            return val, val

                        def create_tax_record(name, amount, tax_type):
                            match = re.search(r'\d+', name)
                            amount = int(match.group()) if match else amount
                            return self.env['account.tax'].create({
                                'name': name,
                                'amount': amount,
                                'type_tax_use': tax_type,
                            }).id

                        account_val, account_name = calculate_tax_value(
                            values['Customer Taxes'])
                        supp_val, supp_name = calculate_tax_value(values['Vendor Taxes'])

                        account_tax = self.env['account.tax'].search(
                            [('name', '=', account_name),
                             ('type_tax_use', '=', 'sale')]).id
                        supp_tax = self.env['account.tax'].search(
                            [('name', '=', supp_name),
                             ('type_tax_use', '=', 'purchase')]).id

                        tax = account_tax if account_tax else create_tax_record(
                            account_name, account_val, 'sale')
                        supplier_tax = supp_tax if supp_tax else create_tax_record(
                            supp_name, supp_val, 'purchase')

                        kay_val_dict = dict(
                            self.env['product.template']._fields[
                                'type'].selection)  # here 'type' is field name
                        for key, val in kay_val_dict.items():
                            if val == values['Product Type']:
                                detailed = key
                        kay_val_dict = dict(
                            self.env['product.template']._fields[
                                'invoice_policy'].selection)  # here 'type' is field name
                        for key, val in kay_val_dict.items():
                            if val == values['Invoicing Policy']:
                                invoicing_type = key
                        if "http://" in values['image'] or "https://" in values[
                            'image']:
                            link = base64.b64encode(requests.get(
                                values['image'].strip()).content).replace(b"\n",
                                                                          b"")
                        elif "/home" in values['image']:
                            if os.path.exists(values['image']):
                                with open(values['image'], 'rb') as file_image:
                                    link = base64.b64encode(file_image.read())

                        if file_reader[0][24] or file_reader[0][25] or \
                                file_reader[0][26]:
                            model = self.env['ir.model']._get_id(
                                'product.template')
                            self.env['ir.model.fields'].create({
                                'model_id': model,
                                'name': file_reader[0][24],
                                'field_description':
                                    file_reader[0][24].split('_')[
                                        2].upper(),
                                'ttype': file_reader[0][24].split('_')[1],
                            })
                            inherit_id = self.env.ref(
                                'product.product_template_only_form_view')
                            arch_base = _('<?xml version="1.0"?>'
                                          '<data>'
                                          '<field name="%s" position="%s">'
                                          '<field name="%s"/>'
                                          '</field>'
                                          '</data>') % (
                                            'type', 'after',
                                            file_reader[0][24])
                            self.env['ir.ui.view'].sudo().create(
                                {'name': 'product.dynamic.fields',
                                 'type': 'form',
                                 'model': 'product.template',
                                 'mode': 'extension',
                                 'inherit_id': inherit_id.id,
                                 'arch_base': arch_base,
                                 'active': True})
                            self.env['ir.model.fields'].create({
                                'model_id': model,
                                'name': file_reader[0][25],
                                'field_description':
                                    file_reader[0][25].split('_')[
                                        2].upper(),
                                'relation': values['Many2many'].split(':')[0],
                                'ttype': file_reader[0][25].split('_')[1],
                            })
                            inherit_id = self.env.ref(
                                'product.product_template_only_form_view')
                            arch_base = _('<?xml version="1.0"?>'
                                          '<data>'
                                          '<field name="%s" position="%s">'
                                          '<field name="%s" widget="%s"/>'
                                          '</field>'
                                          '</data>') % (
                                            'avatax_category_id', 'after',
                                            file_reader[0][25],
                                            'many2many_tags')
                            self.env['ir.ui.view'].sudo().create(
                                {'name': 'product.many2many.fields',
                                 'type': 'form',
                                 'model': 'product.template',
                                 'mode': 'extension',
                                 'inherit_id': inherit_id.id,
                                 'arch_base': arch_base,
                                 'active': True})
                            val = values['Many2many'].split(':')[0]
                            partner = [
                                values['Many2many'].split(':')[1].split(',')]
                            vals_many = []
                            for part in partner[0]:
                                many2many = self.env[val].search(
                                    [('name', '=', part)]).id
                                if many2many:
                                    vals_many.append(many2many)
                                else:
                                    partner = self.env[val].create({
                                        'name': part,
                                    })
                                    vals_many.append(partner)
                            self.env['ir.model.fields'].create({
                                'model_id': model,
                                'name': file_reader[0][26],
                                'field_description':
                                    file_reader[0][26].split('_')[
                                        2].upper(),
                                'relation': values['Many2one'].split(':')[0],
                                'ttype': file_reader[0][26].split('_')[1],
                            })
                            inherit_id = self.env.ref(
                                'product.product_template_only_form_view')
                            arch_base = _('<?xml version="1.0"?>'
                                          '<data>'
                                          '<field name="%s" position="%s">'
                                          '<field name="%s" widget="%s"/>'
                                          '</field>'
                                          '</data>') % (
                                            'categ_id', 'after',
                                            file_reader[0][26],
                                            'many2one_tags')
                            self.env['ir.ui.view'].sudo().create(
                                {'name': 'product.many2one.fields',
                                 'type': 'form',
                                 'model': 'product.template',
                                 'mode': 'extension',
                                 'inherit_id': inherit_id.id,
                                 'arch_base': arch_base,
                                 'active': True})
                            many2one = values['Many2one'].split(':')[0]
                            value = [values['Many2one'].split(':')[1]]
                            vals_one = []
                            for vals in value:
                                many2one_value = self.env[many2one].search(
                                    [('name', '=', vals)]).id
                                if many2one_value:
                                    vals_one.append(many2one_value)
                                else:
                                    value = self.env[many2one].create({
                                        'name': vals,
                                    })
                                    vals_one.append(value)
                            if not values['Internal Reference'] or not values[
                                'Barcode']:
                                raise UserError(
                                    _("File Must  Contain Internal Reference "
                                      "or Barcode of the Product"))
                            if self.method == 'update':
                                vals = {
                                    'default_code': values[
                                        'Internal Reference'] if
                                    values['Internal Reference'] else False,
                                    'name': values['Name'],
                                    'image_1920': link,
                                    'sale_ok': values['Can be sold'],
                                    'purchase_ok': values['Can be Purchased'],
                                    'type': detailed,
                                    'categ_id': pro_category,
                                    'uom_id': uom,
                                    'description': values['Description'],
                                    'barcode': values['Barcode'] if values[
                                        'Barcode'] else False,
                                    'taxes_id': [tax],
                                    'supplier_taxes_id': [supplier_tax],
                                    'description_sale': values[
                                        'Description for customers'],
                                    'invoice_policy': invoicing_type,
                                    'list_price': values['Sales Price'],
                                    'standard_price': values['Cost'],
                                    'weight': values['Weight'],
                                    'volume': values['Volume'],
                                }
                                if link:
                                    vals.update({'image_1920': link})
                                product = self.env[
                                    'product.template'].search(
                                    [('barcode', '=', values['Barcode'])])
                                if len(product):
                                    product.write(vals)
                                else:
                                    product = self.env[
                                        'product.template'].search(
                                        [('default_code', '=',
                                          values['Internal Reference'])])
                                    if product:
                                        product.write(vals)
                                    else:
                                        raise UserError(
                                            _("Please ensure that product "
                                              "having the"
                                              "contains Internal reference or "
                                              "Barcode to with your file."))
                            elif self.method == 'update_product':
                                vals = {
                                    'default_code': values[
                                        'Internal Reference'] if
                                    values['Internal Reference'] else False,
                                    'name': values['Name'],
                                    'image_1920': link,
                                    'sale_ok': values['Can be sold'],
                                    'purchase_ok': values['Can be Purchased'],
                                    'type': detailed,
                                    'categ_id': pro_category,
                                    'uom_id': uom,
                                    'description': values['Description'],
                                    'barcode': values['Barcode'] if values[
                                        'Barcode'] else False,
                                    'taxes_id': [tax],
                                    'supplier_taxes_id': [supplier_tax],
                                    'description_sale': values[
                                        'Description for customers'],
                                    'invoice_policy': invoicing_type,
                                    'lst_price': values['Sales Price'],
                                    'standard_price': values['Cost'],
                                    'weight': values['Weight'],
                                    'volume': values['Volume'],
                                }
                                if link:
                                    vals.update({'image_1920': link})
                                product = self.env[
                                    'product.product'].search(
                                    [('barcode', '=', values['Barcode'])])
                                if len(product):
                                    product.write(vals)
                                else:
                                    product = self.env[
                                        'product.product'].search(
                                        [('default_code', '=',
                                          values['Internal Reference'])])
                                    if product:
                                        product.write(vals)
                                    else:
                                        raise UserError(
                                            _("Please ensure that product "
                                              "having the"
                                              "contains Internal reference or "
                                              "Barcode to with your file."))
                            else:
                                vals = {
                                    'default_code': values[
                                        'Internal Reference'] if
                                    values['Internal Reference'] else False,
                                    'name': values['Name'],
                                    'image_1920': link,
                                    'sale_ok': values['Can be sold'],
                                    'purchase_ok': values['Can be Purchased'],
                                    'type': detailed,
                                    'categ_id': pro_category,
                                    'uom_id': uom,
                                    'description': values['description'],
                                    'barcode': values['Barcode'] if values[
                                        'Barcode'] else False,
                                    'taxes_id': [tax],
                                    'supplier_taxes_id': [supplier_tax],
                                    'description_sale': values[
                                        'Description for customers'],
                                    'invoice_policy': invoicing_type,
                                    'list_price': values['Sales Price'],
                                    'standard_price': values['Cost'],
                                    'weight': values['Weight'],
                                    'volume': values['Volume'],
                                }
                                product = self.env['product.template'].create(vals)
                                product.write({
                                    file_reader[0][24]: values['Char'],
                                    file_reader[0][25]: vals_many[0],
                                    file_reader[0][26]: vals_one[0],
                                })
                            attribute_values = []
                            for val_attribute in values[
                                'Variant Attributes'].split(','):
                                attributes = self.env[
                                    'product.attribute'].search(
                                    [('name', '=', val_attribute)]).id
                                if attributes:
                                    attribute = attributes
                                else:
                                    raise UserError(
                                        _("Please add a valid attribute and "
                                          "their values"))
                                attribute_values.append(
                                    {'attribute': attribute})
                                for value in values['Attribute Values'].split(
                                        ','):
                                    attri_values = self.env[
                                        'product.attribute.value'].search(
                                        [('attribute_id', '=', attribute),
                                         ('name', '=', value)]).ids
                                    if len(attri_values) != 0:
                                        attribute_values.extend(
                                            {attri_values[0]})
                            variant = {}
                            mylist = []
                            for attribute in attribute_values:
                                if isinstance(attribute, dict):
                                    variant = attribute
                                    variant['attribut_value'] = []
                                else:
                                    variant['attribut_value'].extend(
                                        [attribute])
                                if variant in mylist:
                                    pass
                                else:
                                    mylist.append(variant)
                            for list in mylist:
                                val = {
                                    'product_tmpl_id': product.id,
                                    'attribute_id': list['attribute'],
                                    'value_ids': list['attribut_value'],
                                }
                                self.env[
                                    'product.template.attribute.line'].create(
                                    val)
                            return {
                                'type': 'ir.actions.client',
                                'tag': 'reload',
                            }
        except UserError as e:
            raise UserError(str(e))
