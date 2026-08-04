# -*- coding: utf-8 -*-
#############################################################################
#
#    Cybrosys Technologies Pvt. Ltd.
#
#    Copyright (C) 2026-TODAY Cybrosys Technologies(<https://www.cybrosys.com>)
#    Author: Cybrosys Techno Solutions(<https://www.cybrosys.com>)
#
#    You can modify it under the terms of the GNU LESSER
#    GENERAL PUBLIC LICENSE (LGPL v3), Version 3.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU LESSER GENERAL PUBLIC LICENSE (LGPL v3) for more details.
#
#    You should have received a copy of the GNU LESSER GENERAL PUBLIC LICENSE
#    (LGPL v3) along with this program.
#    If not, see <http://www.gnu.org/licenses/>.
#
#############################################################################
import base64
import binascii
import csv
import io
import os
import tempfile
from openpyxl import load_workbook
from odoo import fields, models
from odoo.exceptions import ValidationError


class ImportAttendance(models.TransientModel):
    """ Model for Attendance import wizard. """
    _name = 'import.attendance'
    _description = 'Attendance Import'

    file_type = fields.Selection(
        selection=[('csv', 'CSV File'), ('xls', 'XLS File')],
        string='Select File Type', default='csv',
        help="It helps to select File Type")
    file_upload = fields.Binary(string="Upload File",
                                help="It helps to upload files")

    def action_import_attendance(self):
        """Creating attendance record using uploaded xl/csv files"""
        hr_employee = self.env['hr.employee']
        hr_attendance = self.env['hr.attendance']
        datas = {}

        # -------------------- CSV --------------------
        if self.file_type == 'csv':
            try:
                csv_data = base64.b64decode(self.file_upload)
                data_file = io.StringIO(csv_data.decode("utf-8"))
                data_file.seek(0)
                datas = csv.DictReader(data_file, delimiter=',')
            except Exception:
                raise ValidationError(
                    "File not Valid.\n\nPlease check the type and format "
                    "of the file and try again!"
                )

        # -------------------- XLSX --------------------
        if self.file_type == 'xls':
            try:
                fp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                fp.write(binascii.a2b_base64(self.file_upload))
                fp.flush()
                fp.close()
                workbook = load_workbook(fp.name, data_only=True)
                sheet = workbook.active
            except Exception:
                raise ValidationError(
                    "File not Valid.\n\nPlease check the type and format of the file and try again!"
                )

            rows = list(sheet.iter_rows(values_only=True))
            headers = list(rows[0]) if rows else []
            data = []
            for row in rows[1:]:
                if all(c is None or str(c).strip() == '' for c in row):
                    continue
                data.append({k: v for k, v in zip(headers, row)
                             if k is not None})
            datas = data
            try:
                os.unlink(fp.name)
            except Exception:
                pass

        # -------------------- PROCESS DATA --------------------
        for item in datas:
            vals = {}

            # Employee
            if item.get('Employee'):
                employee = hr_employee.search(
                    [('name', '=', item.get('Employee'))], limit=1
                )
                if employee:
                    vals['employee_id'] = employee.id
                else:
                    raise ValidationError(
                        "There is no employee with that name."
                        "\n\nPlease check and try again!"
                    )

            # Check In
            if item.get('Check In'):
                if self.file_type == 'csv':
                    vals['check_in'] = item.get('Check In')
                else:
                    # openpyxl already gives datetime
                    vals['check_in'] = item.get('Check In')

            # Check Out
            if item.get('Check Out'):
                if self.file_type == 'csv':
                    vals['check_out'] = item.get('Check Out')
                else:
                    vals['check_out'] = item.get('Check Out')

            # Worked Hours
            if item.get('Worked Hours'):
                vals['worked_hours'] = item.get('Worked Hours')

            hr_attendance.create(vals)

        return {
            'effect': {
                'fadeout': 'slow',
                'message': 'Imported Successfully',
                'type': 'rainbow_man',
            }
        }
