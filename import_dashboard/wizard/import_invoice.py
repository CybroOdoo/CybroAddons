# -*- coding: utf-8 -*-
#############################################################################
#
#    Cybrosys Technologies Pvt. Ltd.
#
#    Copyright (C) 2026-TODAY Cybrosys Technologies(<https://www.cybrosys.com>)
#    Author: Cybrosys Techno Solutions(<https://www.cybrosys.com>)
#
#    You can modify it under the terms of the GNU LESSER
#    GENERAL PUBLIC LICENSE (LGPL v3), Version 3.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU LESSER GENERAL PUBLIC LICENSE (LGPL v3) for more details.
#
#    You should have received a copy of the GNU LESSER GENERAL PUBLIC LICENSE
#    (LGPL v3) along with this program.
#    If not, see <http://www.gnu.org/licenses/>.
#
#############################################################################
import base64
import binascii
import csv
import datetime
import io
import os
import re
import tempfile
import xlrd
from openpyxl import load_workbook

from odoo.exceptions import ValidationError
from odoo import fields, models


class ImportInvoice(models.TransientModel):
    """ Model for import invoice """
    _name = 'import.invoice'
    _description = 'Invoice Import'

    file_type = fields.Selection(
        selection=[('csv', 'CSV File'), ('xlsx', 'XLSX File')],
        string='Import File Type', default='csv',
        help="It helps to choose the file type")
    file = fields.Binary(string="File", help="File name")
    update_posted = fields.Boolean(
        string='Update Posted Record?',
        help='If enabled, the records in "Posted" state will converted to draft'
             ' and values are updated. These records will then again be posted'
             ' if "Post Automatically" is activated')
    auto_post = fields.Boolean(string='Post Automatically',
                               help="Post Automatically")
    journal = fields.Selection(
        selection=[('Bank', 'Bank'), ('Cash', 'Cash')], string='Journal',
        default='Bank', help='It helps to choose Journal type')
    order_number = fields.Selection(
        selection=[('from_system', 'From System'), ('from_file', 'From File')],
        string='Number', default='from_file', help="Order number")
    import_product_by = fields.Selection(
        selection=[('name', 'Name'), ('default_code', 'Internal Reference'),
                   ('barcode', 'Barcode')], required=True, default="name",
        string="Import invoice by", help="Product import")
    type = fields.Selection(
        selection=[('out_invoice', 'Invoice'), ('in_invoice', 'Bill'),
                   ('out_refund', 'Credit Note'), ('in_refund', 'Refund')],
        string='Invoicing Type', required=True, help="Invoice type",
        default="out_invoice")

    def action_import_invoice(self):
        """Creating Invoice record using uploaded xl/csv files"""
        account_move = self.env['account.move']
        res_partner = self.env['res.partner']
        res_users = self.env['res.users']
        account_account = self.env['account.account']
        uom_uom = self.env['uom.uom']
        account_tax = self.env['account.tax']
        product_product = self.env['product.product']
        product_attribute = self.env['product.attribute']
        product_attribute_value = self.env['product.attribute.value']
        product_template_attribute_value = self.env[
            'product.template.attribute.value']

        # Support workbook datemode for xlsx -> default None for csv
        workbook_datemode = None

        # ---- Helpers: header normalization + date parsing ----
        def _normalize_header(h):
            if h is None:
                return ''
            return str(h).strip().lower()

        HEADER_MAP = {
            'product': 'Product',
            'product name': 'Product',
            'product_name': 'Product',

            'label': 'Label',
            'description': 'Label',

            'account code': 'Account Code',
            'account_code': 'Account Code',

            'price': 'Price',
            'unit price': 'Price',

            'quantity': 'Quantity',
            'qty': 'Quantity',

            'uom': 'Uom',
            'unit of measure': 'Uom',

            'taxes': 'Taxes',

            'internal reference': 'Internal Reference',
            'default_code': 'Internal Reference',

            'barcode': 'Barcode',

            'number': 'Number',

            'invoice date': 'Invoice Date',
            'inv date': 'Invoice Date',

            'due date': 'Due Date',

            'salesperson': 'Salesperson',

            'payment reference': 'Payment Reference',

            'variant values': 'Variant Values',

            'disc.%': 'Disc.%'
        }

        # Date parsing helper (handles strings, datetime objects, Excel serials)
        def _parse_date(value):
            """
            Return a datetime.date or raise ValueError.
            Accepts:
              - datetime.date / datetime.datetime
              - Excel serial numbers (float/int) if workbook_datemode available
              - Strings in several common formats
            """
            if value in (None, ''):
                return None
            # Already a date/datetime
            if isinstance(value, datetime.date) and not isinstance(value, datetime.datetime):
                return value
            if isinstance(value, datetime.datetime):
                return value.date()
            # Excel numeric date (from xlsx): ctype gave us a float in sheet.row_values
            if isinstance(value, (float, int)):
                if workbook_datemode is not None:
                    try:
                        dt = xlrd.xldate_as_datetime(value, workbook_datemode)
                        return dt.date()
                    except Exception:
                        raise ValueError("Invalid Excel date serial: %s" % value)
                else:
                    # No workbook datemode — treat numeric as invalid
                    raise ValueError("Numeric date encountered but workbook datemode is unknown.")
            # Try string formats
            if isinstance(value, str):
                value = value.strip()
                # Some CSV exports may contain timestamps; remove time portion if present
                if ' ' in value and value.count(':') >= 1:
                    # keep only date part
                    value = value.split(' ')[0].strip()
                # formats to attempt (order matters)
                fmts = ['%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d.%m.%Y', '%Y/%m/%d', '%m-%d-%Y']
                for f in fmts:
                    try:
                        dt = datetime.datetime.strptime(value, f)
                        return dt.date()
                    except Exception:
                        continue
            raise ValueError("Unrecognized date format: %s" % value)

        # ---- Read file into normalized items list ----
        items = []
        if self.file_type == 'csv':
            try:
                csv_data = base64.b64decode(self.file)
                data_file = io.StringIO(csv_data.decode("utf-8"))
                data_file.seek(0)
                reader = csv.DictReader(data_file)
            except Exception:
                raise ValidationError(
                    "File not Valid.\n\nPlease check the type and format "
                    "of the file and try again!")
            for raw in reader:
                row = {}
                for raw_key, raw_val in raw.items():
                    nk = _normalize_header(raw_key)
                    row_key = HEADER_MAP.get(nk, raw_key.strip() if raw_key else raw_key)
                    row[row_key] = raw_val
                items.append(row)

        if self.file_type == 'xlsx':
            try:
                fp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                fp.write(binascii.a2b_base64(self.file))
                fp.flush()
                fp.close()
                workbook = load_workbook(fp.name, data_only=True)
                sheet = workbook.active
            except Exception:
                raise ValidationError(
                    "File not Valid.\n\nPlease check the "
                    "type and format of the file and try again!")
            rows = list(sheet.iter_rows(values_only=True))
            raw_headers = list(rows[0]) if rows else []
            norm_headers = [_normalize_header(h) for h in raw_headers]
            headers = []
            for nh, rh in zip(norm_headers, raw_headers):
                headers.append(HEADER_MAP.get(
                    nh, rh.strip() if isinstance(rh, str) else rh))
            for row_vals in rows[1:]:
                if all(c is None or str(c).strip() == '' for c in row_vals):
                    continue
                items.append({k: v for k, v in zip(headers, row_vals)
                              if k is not None})
            try:
                os.unlink(fp.name)
            except Exception:
                pass

        # ---- Main import loop ----
        row = 0
        imported = 0
        confirmed = 0
        imported_invoices = []
        error_msg = ""
        partner_added_msg = ""
        warning_msg = ""
        if items:
            for item in items:
                row += 1
                vals = {}
                row_not_import_msg = "\n❌Row {rn} not imported.".format(rn=row)
                import_error_msg = ""
                missing_fields_msg = ""
                fields_msg = "\n\tMissing required field(s):"
                partner_msg = "\n🆕New Partner(s) added:"
                vals['move_type'] = self.type

                # Partner
                partner_name = item.get('Partner')
                if partner_name:
                    partner = res_partner.search([('name', '=', partner_name)])
                    if not partner:
                        partner = res_partner.create({'name': partner_name})
                        vals['partner_id'] = partner.id
                        if partner_added_msg:
                            partner_added_msg += ("\n\t\trow {rn}: {partner}").format(
                                rn=row, partner=partner_name)
                        else:
                            partner_added_msg += (partner_msg + "\n\t\trow {rn}: "
                                                                "\"{partner}\"").format(rn=row, partner=partner_name)
                    elif len(partner) > 1:
                        import_error_msg += row_not_import_msg + (
                                "\n\t\t⚠ Multiple Partners with name (%s) found!"
                                % partner_name)
                    else:
                        vals['partner_id'] = partner.id
                else:
                    if missing_fields_msg:
                        missing_fields_msg += "\n\t\t❗ \"Partner\""
                    else:
                        missing_fields_msg += (fields_msg + "\n\t\t❗ \"Partner\"")

                # Missing fields accumulation
                if import_error_msg:
                    import_error_msg += missing_fields_msg
                elif missing_fields_msg:
                    import_error_msg += (row_not_import_msg + missing_fields_msg)

                # Payment reference
                if item.get('Payment Reference'):
                    vals['payment_reference'] = item['Payment Reference']

                # Invoice date
                if item.get('Invoice Date'):
                    date = item['Invoice Date']
                    try:
                        invoice_date = _parse_date(date)
                        if invoice_date:
                            vals['invoice_date'] = invoice_date
                    except ValueError as e:
                        import_error_msg += row_not_import_msg + (
                                "\n\t\t⚠ Please check the date format of Invoice Date. (%s)" % e)

                # Due date
                if item.get('Due Date'):
                    date = item['Due Date']
                    try:
                        due_date = _parse_date(date)
                        if due_date:
                            vals['invoice_date_due'] = due_date
                    except ValueError as e:
                        import_error_msg += row_not_import_msg + (
                                "\n\t\t⚠ Please check the date format of Due Date. (%s)" % e)

                # Salesperson
                if item.get('Salesperson'):
                    sales_person = res_users.search([('name', '=', item['Salesperson'])], limit=1)
                    if sales_person:
                        vals['invoice_user_id'] = sales_person.id

                # If any import-level error so far, skip this row
                if import_error_msg:
                    error_msg += import_error_msg
                    continue

                # Find or create invoice
                invoice = account_move.search(
                    [('name', '=', item.get('Number')), ('move_type', '=', vals['move_type'])])
                if invoice:
                    if len(invoice) > 1:
                        error_msg += row_not_import_msg + (
                                "\n\t⚠ Multiple invoice with same Number(%s) found!" % item.get('Number'))
                        continue
                    if vals:
                        if self.update_posted and invoice.state == 'posted':
                            invoice.button_draft()
                            invoice.write(vals)
                        elif invoice.state == 'draft':
                            invoice.write(vals)
                else:
                    if self.order_number == 'from_system':
                        invoice = account_move.create(vals)
                    elif self.order_number == 'from_file':
                        if item.get('Number'):
                            vals['name'] = item['Number']
                            invoice = account_move.create(vals)
                        else:
                            error_msg += (row_not_import_msg + fields_msg + "\n\t\t\"Number\"")
                            continue

                # ---- Prepare invoice line values ----
                line_vals = {}
                pro_vals = {}

                # Description fallback: Label > Product > Internal Reference > Barcode
                label_value = item.get('Label') or False
                product_value = item.get('Product') or item.get('Internal Reference') or item.get('Barcode') or False

                if label_value:
                    line_vals['name'] = label_value
                elif product_value:
                    line_vals['name'] = product_value
                else:
                    has_amount_info = bool(item.get('Price')) or bool(item.get('Quantity'))
                    if has_amount_info:
                        line_vals['name'] = "Imported Line (row %d)" % row
                        warning_msg += "\n◼ Row %d: no Product/Label — created generic line description." % row
                    else:
                        import_error_msg += row_not_import_msg + "\n\t⚠ Product and Label missing in file!"
                        error_msg += import_error_msg
                        continue

                # Safe account lookup
                if item.get('Account Code'):
                    raw_code = str(item['Account Code']).strip()
                    account = account_account.search([('code', '=', raw_code)], limit=1)
                    if not account:
                        try:
                            num = int(float(raw_code))
                            account = account_account.search([('code', '=', str(num))], limit=1)
                        except Exception:
                            account = False
                    if account:
                        line_vals['account_id'] = account.id

                # Quantity
                if item.get('Quantity'):
                    # try numeric conversion
                    try:
                        line_vals['quantity'] = float(item.get('Quantity'))
                    except Exception:
                        line_vals['quantity'] = item.get('Quantity')

                # UoM
                if item.get('Uom'):
                    uom = uom_uom.search([('name', '=', item['Uom'])], limit=1)
                    if uom:
                        pro_vals['uom_id'] = line_vals['product_uom_id'] = uom.id

                # Price
                if item.get('Price'):
                    try:
                        price_val = float(item.get('Price'))
                        pro_vals['lst_price'] = line_vals['price_unit'] = price_val
                    except Exception:
                        pro_vals['lst_price'] = line_vals['price_unit'] = item.get('Price')

                # Discount
                if item.get('Disc.%'):
                    try:
                        line_vals['discount'] = float(item.get('Disc.%'))
                    except Exception:
                        line_vals['discount'] = item.get('Disc.%')

                # Taxes
                if item.get('Taxes'):
                    tax_name = item['Taxes']
                    m = re.findall(r"(\d+)%", tax_name)
                    tax_amount = m[0] if m else 0.0
                    tax = account_tax.search([('name', '=', tax_name), ('type_tax_use', '=', 'sale')], limit=1)
                    if not tax:
                        tax = account_tax.create({
                            'name': tax_name,
                            'type_tax_use': 'sale',
                            'amount': float(tax_amount) if tax_amount else 0.0
                        })
                    pro_vals['taxes_id'] = line_vals['tax_ids'] = [tax.id]

                # Product fields
                if item.get('Product'):
                    pro_vals['name'] = item['Product']
                if item.get('Internal Reference'):
                    pro_vals['default_code'] = item['Internal Reference']
                if item.get('Barcode'):
                    pro_vals['barcode'] = item['Barcode']

                # --- product resolution depending on import_product_by ---
                product = False
                if self.import_product_by == 'name':
                    if item.get('Product'):
                        product = product_product.search([('name', '=', item['Product'])])
                        if not product:
                            # create product with provided pro_vals
                            product = product_product.create(pro_vals)
                        if len(product) > 1:
                            # handle variants if provided
                            if item.get('Variant Values'):
                                pro_tmpl_id = product.mapped('product_tmpl_id')
                                if len(pro_tmpl_id) > 1:
                                    error_msg += row_not_import_msg + (
                                            "\n\t⚠ Multiple Product records are linked with the product variant \"%s\"." %
                                            item['Product'])
                                    continue
                                variant_values = item['Variant Values'].split(',')
                                variant_value_ids = []
                                for var in variant_values:
                                    k_v = var.partition(":")
                                    attr = k_v[0].strip()
                                    attr_val = k_v[2].strip()
                                    var_attr_ids = product_attribute.search([('name', '=', attr)]).ids
                                    var_attr_val_ids = product_attribute_value.search(
                                        [('name', '=', attr_val), ('attribute_id', 'in', var_attr_ids)]).ids
                                    if var_attr_val_ids:
                                        pro_temp_attr_val_id = product_template_attribute_value.search(
                                            [('product_attribute_value_id', 'in', var_attr_val_ids),
                                             ('product_tmpl_id', '=', pro_tmpl_id.id)]).id
                                        variant_value_ids += [pro_temp_attr_val_id]
                                if variant_value_ids:
                                    product = product.filtered(
                                        lambda p: p.product_template_variant_value_ids.ids == variant_value_ids)
                                else:
                                    error_msg += row_not_import_msg + (
                                            "\n\t⚠ Product variant with variant values \"%s\" not found." % (
                                    item['Variant Values']))
                                    continue
                                if len(product) != 1:
                                    error_msg += row_not_import_msg + (
                                            "\n\t⚠ Multiple variants with same Variant Values \"%s\" found." % (
                                    item['Variant Values']))
                                    continue
                            else:
                                error_msg += row_not_import_msg + (
                                        "\n\t⚠ Multiple Products with same Name \"%s\" found. Provide unique product Variant Values." % (
                                item['Product']))
                                continue
                    else:
                        error_msg += row_not_import_msg + ("\n\tProduct name missing in file!")
                        continue

                if self.import_product_by == 'default_code':
                    if item.get('Internal Reference'):
                        product = product_product.search([('default_code', '=', item['Internal Reference'])])
                        if not product:
                            if not item.get('Product'):
                                warning_msg += (
                                            "\n◼ A Product is created with \"Internal Reference\" as product name since \"Product\" name is missing in file. (row %d)" % row)
                                pro_vals['name'] = item['Internal Reference']
                            product = product_product.create(pro_vals)
                        if len(product) > 1:
                            error_msg += row_not_import_msg + (
                                    "\n\t⚠ Multiple Products with same Internal Reference(%s) found!" % item[
                                'Internal Reference'])
                            continue
                    else:
                        error_msg += row_not_import_msg + ("\n\tInternal Reference missing in file!")
                        continue

                if self.import_product_by == 'barcode':
                    if item.get('Barcode'):
                        product = product_product.search([('barcode', '=', item['Barcode'])])
                        if not product:
                            if not item.get('Product'):
                                warning_msg += (
                                            "\n◼ No value under \"Product\" at row %d, thus added \"Barcode\" as product name" % row)
                                pro_vals['name'] = item['Barcode']
                            product = product_product.create(pro_vals)
                        if len(product) > 1:
                            error_msg += row_not_import_msg + (
                                        "\n\tOther Product(s) with same Barcode (%s) found!" % item['Barcode'])
                            continue
                    else:
                        error_msg += row_not_import_msg + ("\n\tBarcode missing in file!")
                        continue

                # Attach product and create line
                if product:
                    line_vals['product_id'] = product.id

                # Ensure minimal required fields for account.move.line creation:
                # if account_id is not set but product has income account, Odoo may compute it.
                # Create the line on invoice
                invoice.write({'invoice_line_ids': [(0, 0, line_vals)]})

                imported += 1
                imported_invoices += [invoice]

                # Auto-post if requested (do after collecting imported_invoices)
                if self.auto_post and imported_invoices:
                    for inv in imported_invoices:
                        inv.action_post()
                        confirmed += 1

            # End for items

            # Report errors if any
            if error_msg:
                error_msg = "\n\nWARNING" + error_msg
                error_message = self.env['import.message'].create({'message': error_msg})
                return {
                    'name': 'Error!',
                    'type': 'ir.actions.act_window',
                    'view_mode': 'form',
                    'res_model': 'import.message',
                    'res_id': error_message.id,
                    'target': 'new'
                }

            # Success message
            msg = (("Imported %d records.\nPosted %d records" % (imported,
                                                                 confirmed)) + partner_added_msg + warning_msg)
            message = self.env['import.message'].create({'message': msg})
            if message:
                return {
                    'effect': {
                        'fadeout': 'slow',
                        'message': msg,
                        'type': 'rainbow_man',
                    }
                }
