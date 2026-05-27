# -*- coding: utf-8 -*-
"""Lots Attachment."""
###############################################################################
#
#    Cybrosys Technologies Pvt. Ltd.
#
#    Copyright (C) 2026-TODAY Cybrosys Technologies(<https://www.cybrosys.com>)
#    Author: Cybrosys Techno Solutions (odoo@cybrosys.com)
#
#    You can modify it under the terms of the GNU LESSER
#    GENERAL PUBLIC LICENSE (LGPL v3), Version 3.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU LESSER GENERAL PUBLIC LICENSE (LGPL v3) for more details.
#
#    You should have received a copy of the GNU LESSER GENERAL PUBLIC
#    LICENSE (LGPL v3) along with this program.
#    If not, see <http://www.gnu.org/licenses/>.
#
###############################################################################
import base64
from io import BytesIO

import openpyxl

from odoo import _, fields, models
from odoo.exceptions import UserError, ValidationError


class LotsAttachment(models.TransientModel):
    """ Class for lots wizard """
    _name = 'lot.attachment'
    _description = "Lots Attachment"

    picking_id = fields.Many2one('stock.picking',
                                 string="Stock Picking",
                                 help="Parent picking")
    product_id = fields.Many2one('product.product',
                                 string="Product",
                                 help="Current product")
    demanded_quantity = fields.Float(string="Quantity",
                                     help="Product quantity demanded")
    type = fields.Selection(string="Lots Type",
                            selection=[('lot', 'Lot'), ('serial', 'Serial')],
                            help="Choose a lot/serial")
    move_id = fields.Many2one('stock.move')
    attachment = fields.Binary(string="Upload", attachment=True)
    attachment_name = fields.Char(string="Attachment Name",
                                  help="Attachment file name")

    def action_import_lot(self):
        """ Importing lots """
        current_move_id = self.env['stock.move'].browse(self.move_id.id)
        if self.attachment:
            wb = openpyxl.load_workbook(
                filename=BytesIO(base64.b64decode(self.attachment)),
                read_only=True) if self.attachment else ""
            ws = wb.active
            # Check if product exists in the sheet
            product_found = any(
                record[1] == current_move_id.product_id.display_name for
                record in ws.iter_rows(min_row=2, values_only=True))
            if not product_found:
                raise UserError(
                    _('The product "%s" does not exist in the sheet.') %
                    current_move_id.product_id.display_name)
            # Check if lot name already exists in move line ids
            if (current_move_id.move_line_ids and
                    any(record[0] in set(current_move_id.move_line_ids.mapped('lot_name')) for
                    record in ws.iter_rows(min_row=2, values_only=True))):
                raise UserError(
                    _('This Lot name already exists in the move line.'))
            # Calculate total sheet quantity for the product
            total_sheet_quantity = sum(
                record[2] for record in ws.iter_rows
                (min_row=2, values_only=True) if
                record[1] == current_move_id.product_id.display_name)
            # Check if total sheet quantity exceeds demand quantity of the
            # product
            if total_sheet_quantity > current_move_id.product_uom_qty:
                raise UserError(
                    _('Total quantity in the sheet exceeds the demand '
                      'quantity of the product. Please adjust the quantities '
                      'in the sheet.'))
            # Prepare move line values to be written
            vals_list = []
            for record in ws.iter_rows(min_row=2, values_only=True):
                lot_name, product_name, quantity = record
                if product_name == current_move_id.product_id.display_name:
                    lot_name_str = str(lot_name) if lot_name else False

                    # Search for existing lot
                    existing_lot = self.env['stock.lot'].search([
                        ('name', '=', lot_name_str),
                        ('product_id', '=', current_move_id.product_id.id),
                        ('company_id', '=', current_move_id.company_id.id)
                    ], limit=1)

                    move_line_vals = {
                        'quantity': quantity,
                        'move_id': current_move_id.id,
                        'product_id': current_move_id.product_id.id,
                        'picking_id': current_move_id.picking_id.id,
                        'location_id': current_move_id.location_id.id,
                        'location_dest_id': current_move_id.location_dest_id.id,
                    }

                    if existing_lot:
                        move_line_vals['lot_id'] = existing_lot.id
                    else:
                        move_line_vals['lot_name'] = lot_name_str

                    vals_list.append((0, 0, move_line_vals))
            # Write move line values
            current_move_id.move_line_ids.unlink()
            current_move_id.write({'move_line_ids': vals_list})
            return {
                'type': 'ir.actions.client',
                'tag': 'reload',
            }
        raise ValidationError(
            _('Check whether you upload the document'))

    def action_download_sample(self):
        """ For downloading a sample excel file """
        return {
            'type': 'ir.actions.act_url',
            'url': '/download/excel',
            'target': 'self',
            'file_name': 'my_excel_file.xlsx'
        }
